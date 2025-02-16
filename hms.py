import pandas
import os
import openpyxl
z = os.path.exists("HMSEXCEL.xlsx")
if z:
    wb = openpyxl.load_workbook("HMSEXCEL.xlsx")
else:
    wb = openpyxl.Workbook()
    wb.save("HMSEXCEL.xlsx")
    sh = wb.active
    sh.title = "doctordetails"
    wb["doctordetails"].append(["DOCTOR_ID", "DOCTOR_NAME", "SPECILIZATION", "EXPERIENCE"])
    wb.create_sheet(title="coworkersdetails")
    wb["coworkersdetails"].append(["WORKER_ID", "NAME", "QUALIFICATION", "POSITION"])
    wb.create_sheet(title="paitentdetails")
    wb["paitentdetails"].append(["PATIENT_TOKEN_NUMBER", "PATIENT_NAME", "GENDER", "AGE", "ADDRESS"])
    wb.save("HMSEXCEL.xlsx")


sheet1 = wb["doctordetails"]
sheet2 = wb["coworkersdetails"]
sheet3 = wb["paitentdetails"]


def display():
    try:
        xls = pandas.ExcelFile("HMSEXCEL.xlsx")
        data1 = pandas.read_excel(xls, "doctordetails")
        data2 = pandas.read_excel(xls, "coworkersdetails")
        data3 = pandas.read_excel(xls, "paitentdetails")
        print("\t\t1. Doctors Details\n\t\t2. Co-worker Details\n\t\t3. Patient Details")
        a = int(input("enter your choice: "))
        if a == 1:
            if data1.empty:
                print("-----DATA IS EMPTY-----")
            else:
                print(data1)
        elif a == 2:
            if data2.empty:
                print("-----DATA IS EMPTY-----")
            else:
                print(data2)
        elif a == 3:
            if data3.empty:
                print("-----DATA IS EMPTY-----")
            else:
                print(data3)
        else:
            print("Sorry, Your entered wrong choice")
    except ValueError:
        print("Oops, You entered letters/symbols\n Please Try again.....")


def enter():
    try:
        print("\t\t1. Doctors Details\n\t\t2. Co-worker Details\n\t\t3. Patient Details")
        c = int(input("Enter your Choice:"))
        if c == 1:
            dd = int(input("Enter the Doctor Id\n\t").strip())
            di = input("Enter the Doctor Name\n\t").strip()
            dj = input("Enter the Specilization\n\t").strip()
            dk = input("Enter your Experince\n\t").strip()
            sheet1.append([dd, di, dj, dk])
            wb.save("HMSEXCEL.xlsx")
        elif c == 2:
            cd = int(input("Enter the Worker Id\n\t").strip())
            cn = input("Enter your Name\n\t").strip()
            cq = input("Enter Your Qualificationn\n\t").strip()
            cp = input("Enter your position\n\t").strip()
            sheet2.append([cd, cn, cq, cp])
            wb.save("HMSEXCEL.xlsx")
        elif c == 3:
            pi = int(input("Enter Paitent Token Number\n\t").strip())
            pn = input("enter the paitent name\n\t").strip()
            pa = input("enter the paitent age\n\t").strip()
            pg = input("enter the Gender\n\t").strip()
            pd = input("enter the address\n\t").strip()
            sheet3.append([pi, pn, pg, pa, pd])
            wb.save("HMSEXCEL.xlsx")
        else:
            print("Sorry, Your entered wrong choice")
    except ValueError:
        print("Oops, You entered letters/symbols\n Please Try again.....")


def delete():
    print("\t\t1. Doctors Details\n\t\t2. Co-worker Details\n\t\t3. Patient Details")
    try:
        d = int(input("enter your choice: "))
        if d == 1:
            rem = int(input("Enter the Doctor id: "))
            h = 0
            for cell in sheet1["A"]:
                if cell.value == rem:
                    sheet1.delete_rows(cell.row, 1)
                    print("Sucessfully deleted")
                    h = 1
            if h == 0:
                print("DOCTOR ID IS NOT FOUND")
        elif d == 2:
            rem = int(input("Enter the Worker id: "))
            h = 0
            for cell in sheet2["A"]:
                if cell.value == rem:
                    sheet1.delete_rows(cell.row, 1)
                    print("Sucessfully deleted")
                    h = 1
            if h == 0:
                print("WORKER ID IS NOT FOUND")
        elif d == 3:
            rem = int(input("Enter the Paitent token number: "))
            h = 0
            for cell in sheet3["A"]:
                if cell.value == rem:
                    sheet1.delete_rows(cell.row, 1)
                    print("Sucessfully deleted")
                    h = 1
            if h == 0:
                print("PATIENT TOKEN NUMBER IS NOT FOUND")
        else:
            print("Sorry, Your entered wrong choice")
        wb.save("HMSEXCEL.xlsx")
    except ValueError:
        print("Oops, You entered letters/symbols\n Please Try again.....")


print("::::::::::::::::::::HOSPITAL MANAGEMENT SYSTEM::::::::::::::::::::")
e = 1
while e != 0:
    print("""
                1. Display the details
                2. Add a new member
                3. Delete a member
                4. Make an exit
                                         """)
    try:
        b = int(input("Enter your Choice:"))
        if b == 1:
            display()
        elif b == 2:
            enter()
        elif b == 3:
            delete()
        elif b == 4:
            e = 0
        else:
            print("You entered wrong choice, Please Try Again")
    except ValueError:
        print("Oops, You entered letters/symbols\n Please Try again.....")
